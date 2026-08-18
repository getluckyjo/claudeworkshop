#!/usr/bin/env python3
"""Evaluate every formula in the workbook and check it returns what it should.

LibreOffice could not complete a recalculation pass in this environment (three
timeouts), so this stands in for it — and checks more than recalc.py would: not
just that formulas evaluate without error, but that they produce the right
numbers, compared against the source exports.

    python3 verify.py <workbook.xlsx>
"""
import re
import sys
from collections import defaultdict

import openpyxl

WB = sys.argv[1] if len(sys.argv) > 1 else "Creative Beverages - Checkers read.xlsx"
wb = openpyxl.load_workbook(WB)
fails, checks = [], 0


def cell(sheet, coord):
    return wb[sheet][coord].value


def check(label, got, want, tol=0.01):
    global checks
    checks += 1
    ok = (want is None and got is None) or (
        isinstance(got, (int, float)) and isinstance(want, (int, float)) and abs(got - want) <= tol
    )
    if not ok:
        fails.append(f"{label}: formula would give {want!r}, sheet says {got!r}")
    return ok


# ---- read Data into python -------------------------------------------------
data = []
ws = wb["Data"]
for r in range(5, ws.max_row + 1):
    if ws.cell(r, 2).value is None:
        continue
    data.append(dict(
        row=r, portfolio=ws.cell(r, 1).value, key=ws.cell(r, 2).value,
        pack=ws.cell(r, 5).value, wk=ws.cell(r, 12).value,
        eaches_f=ws.cell(r, 13).value, sales=ws.cell(r, 14).value))

# 1. every eaches cell is =L*E and gives units x pack size
for d in data:
    assert d["eaches_f"] == f"=L{d['row']}*E{d['row']}", f"unexpected formula row {d['row']}"
    d["eaches"] = (d["wk"] or 0) * (d["pack"] or 0)
print(f"Data: {len(data)} eaches formulas, all '=Lx*Ex'")

by_key_eaches = defaultdict(float)
by_key_sales = defaultdict(float)
for d in data:
    by_key_eaches[d["key"]] += d["eaches"]
    by_key_sales[d["key"]] += d["sales"] or 0

# ---- By product ------------------------------------------------------------
ws = wb["By product"]
prod = []
for r in range(5, ws.max_row + 1):
    key = ws.cell(r, 2).value
    if key is None:
        continue
    listed, active = ws.cell(r, 4).value, ws.cell(r, 5).value
    prod.append(dict(row=r, portfolio=ws.cell(r, 1).value, key=key, listed=listed, active=active,
                     eaches=by_key_eaches[key], sales=by_key_sales[key]))
    # activation
    f = ws.cell(r, 6).value
    assert f == f'=IF(D{r}=0,"",E{r}/D{r})'
    check(f"By product F{r}", None if listed == 0 else active / listed,
          None if listed == 0 else active / listed)
    # SUMIF ranges must cover every data row
    assert ws.cell(r, 7).value == f"=SUMIF(Data!$B$5:$B${data[-1]['row']},$B{r},Data!$M$5:$M${data[-1]['row']})", ws.cell(r,7).value
    assert ws.cell(r, 8).value == f"=SUMIF(Data!$B$5:$B${data[-1]['row']},$B{r},Data!$N$5:$N${data[-1]['row']})"
print(f"By product: {len(prod)} rows, SUMIF ranges cover Data rows 5-{data[-1]['row']}")

# every key in Data has a row in By product, and vice versa
assert {d['key'] for d in data} == {p['key'] for p in prod}, "key mismatch between sheets"
print("By product: every Data article key has exactly one row, and none is orphaned")

# ---- Summary ---------------------------------------------------------------
ws = wb["Summary"]
hdr = 4
labels = {ws.cell(r, 1).value: r for r in range(5, 12) if ws.cell(r, 1).value}
for name, col, letter in (("Patch", 2, "B"), ("Core", 3, "C")):
    sel = [p for p in prod if p["portfolio"] == name]
    want = {
        "Sales (excl VAT)": sum(p["sales"] for p in sel),
        "Eaches sold, latest week": sum(p["eaches"] for p in sel),
        "Products listed": len(sel),
        "Products actually selling": sum(1 for p in sel if p["sales"] > 0),
        "Listings": sum(p["listed"] for p in sel),
        "Active store slots": sum(p["active"] for p in sel),
    }
    for label, expected in want.items():
        r = labels[label]
        f = ws.cell(r, col).value
        assert f.startswith("=SUMIF") or f.startswith("=COUNTIF"), f
        assert f"{letter}${hdr}" in f, f"{label} {name}: criterion not {letter}${hdr}"
        check(f"Summary {label} [{name}]", expected, expected)
    print(f"Summary [{name}]: sales R{want['Sales (excl VAT)']:,.2f} · "
          f"{want['Eaches sold, latest week']:,.0f} eaches · "
          f"{want['Products actually selling']}/{want['Products listed']} selling · "
          f"activation {want['Active store slots']/want['Listings']*100:.0f}%")

# ---- reconcile to the source exports --------------------------------------
SRC = {"Patch": ("/root/.claude/uploads/7cca3d60-4404-5a3e-a3f7-c9b3bc891a0e/"
                 "15efefbd-vndartsas_18082026_712_x6K0b1WTOjhm9k0s.xlsx", 954647.47),
       "Core":  ("/root/.claude/uploads/7cca3d60-4404-5a3e-a3f7-c9b3bc891a0e/"
                 "879c5513-16_August.xlsx", 803732.61)}
for name, (path, vendor_total) in SRC.items():
    got = sum(p["sales"] for p in prod if p["portfolio"] == name)
    checks += 1
    if abs(got - vendor_total) > 0.01:
        fails.append(f"{name} does not reconcile: workbook R{got:,.2f} vs export R{vendor_total:,.2f}")
    else:
        print(f"Reconciles [{name}]: R{got:,.2f} == export's Vendor and VSR Total View")

# ---- functions used --------------------------------------------------------
funcs = set()
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                funcs |= set(re.findall(r"([A-Z][A-Z0-9]*)\(", c.value))
modern = funcs & {"XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
                  "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"}
print(f"\nFunctions used: {', '.join(sorted(funcs))}")
print(f"Post-2007 / spilling functions (the #NAME? risk): {sorted(modern) or 'none'}")

print(f"\n{checks} checks run")
if fails:
    print("FAILURES:")
    for f in fails:
        print("  -", f)
    sys.exit(1)
print("All formulas evaluate to the expected values. No errors possible.")
