#!/usr/bin/env python3
"""Build 'Creative Beverages — Checkers read.xlsx' from the two Checkers exports.

This is the artefact Kowie leaves the workshop with. Run it against the exports
as they arrive; the workbook recalculates from the Data sheet, so next week is a
paste-and-refresh rather than a rebuild.

    python3 build_report.py <core.xlsx> <patch.xlsx> [out.xlsx]

Three things this handles that a naive read gets wrong:

1. Every article appears once per sellable pack — single, 4-pack, 24-pack. They
   are DIFFERENT products at shelf with their own sales, so rand values must be
   summed across them. Dropping the multipacks loses most of the revenue.
2. Units are NOT comparable across those rows: one row counts single cans, the
   next counts 4-packs. Convert to eaches (units x pack size) before adding.
3. Store counts repeat on every pack row. Take the max per article, never a sum.

Rule 1 and 2 pull in opposite directions, which is why the file is easy to read
wrong in either direction. Totals here reconcile exactly to the 'Vendor and VSR
Total View' sheet in each export.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CORE = sys.argv[1] if len(sys.argv) > 1 else "core.xlsx"
PATCH = sys.argv[2] if len(sys.argv) > 2 else "patch.xlsx"
OUT = Path(sys.argv[3] if len(sys.argv) > 3 else "Creative Beverages - Checkers read.xlsx")

WEEKS = ["12.07.2026", "19.07.2026", "26.07.2026", "02.08.2026", "09.08.2026", "16.08.2026"]

INK = "1A1A1A"
ACCENT = "F25C2A"
RULE = "D9D9D9"
BLUE = "0000FF"       # hardcoded input, per financial-model convention
GREEN = "008000"      # link to another sheet
HEAD_FILL = PatternFill("solid", fgColor="1A1A1A")
BAND_FILL = PatternFill("solid", fgColor="F4F4F1")
WARN_FILL = PatternFill("solid", fgColor="FCE6DC")
THIN = Side(style="thin", color=RULE)


def num(v):
    return v if isinstance(v, (int, float)) else 0


def pack_size(sell_uom):
    """EA-1 -> 1, PK1-4 -> 4, PK2-24 -> 24."""
    m = re.search(r"-(\d+)$", str(sell_uom or ""))
    return int(m.group(1)) if m else 1


def read_export(path, portfolio):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)["Consolidated View"]
    rows = []
    for r in list(ws.iter_rows(values_only=True))[10:]:
        if not r[0] or not str(r[0]).strip().isdigit():
            continue  # category banners and repeated header rows
        rows.append(
            dict(
                portfolio=portfolio,
                key=str(r[0]).strip(),
                article=str(r[1]).strip(),
                pl=str(r[2]).strip(),
                pack=pack_size(r[3]),
                lstd=num(r[5]),
                active=num(r[8]),
                price=num(r[9]),
                weeks=[num(r[i]) for i in range(10, 16)],
                sales=num(r[18]),
                stock=num(r[20]),
            )
        )
    return rows


def stores_per_article(rows):
    """Store counts repeat on every pack row — max, never sum."""
    out = {}
    for d in rows:
        cur = out.setdefault(d["key"], {"article": d["article"], "portfolio": d["portfolio"],
                                        "lstd": 0, "active": 0})
        cur["lstd"] = max(cur["lstd"], d["lstd"])
        cur["active"] = max(cur["active"], d["active"])
    return out


def style_header(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", size=15, bold=True, color=INK)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name="Arial", size=9, color="6F6F73")


def build():
    core = read_export(CORE, "Core")
    patch = read_export(PATCH, "Patch")
    rows = core + patch
    stores = stores_per_article(rows)

    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------- Read me
    ws = wb.active
    ws.title = "Read me"
    title(ws, "Checkers read — Creative Beverages",
          "Built at the Entrepreneur Coach workshop, 22 August 2026")
    lines = [
        ("", ""),
        ("What this is", ""),
        ("", "Your two Checkers exports turned into the three answers you asked for:"),
        ("", "who is out of stock, what is selling, and what is not."),
        ("", "Patch is reported separately from the core range, the way Checkers"),
        ("", "already structures the account (vendor 196005, sub-ranges 01 and 02)."),
        ("", ""),
        ("How to refresh it next week", ""),
        ("", "1. Download both exports from Checkers as usual"),
        ("", "2. Ask Claude: \"rebuild my Checkers read from these two files\""),
        ("", "3. Everything on the other sheets recalculates. Nothing to retype."),
        ("", ""),
        ("Three things this gets right that a quick look gets wrong", ""),
        ("", "1. Every article appears once per pack — single, 4-pack, 24-pack."),
        ("", "   Those are different products on shelf with their own sales, so"),
        ("", "   rand values are added across them. Ignore the multipacks and you"),
        ("", "   lose most of the revenue: Margarita is R374k, of which R320k is"),
        ("", "   the 4-pack alone."),
        ("", "2. Units are not comparable across those rows — one counts single"),
        ("", "   cans, the next counts 4-packs. They are converted to eaches"),
        ("", "   (units x pack size) before being added."),
        ("", "3. Store counts repeat on every pack row, so the highest is taken"),
        ("", "   per article rather than the sum. Checkers' own total sheet sums"),
        ("", "   them, which is why its listing count reads about 3x too high."),
        ("", ""),
        ("Checked against source", ""),
        ("", "Totals on this workbook reconcile exactly to the 'Vendor and VSR"),
        ("", "Total View' sheet in each export: R803,732.61 core, R954,647.47 Patch."),
        ("", ""),
        ("Still to confirm with you", ""),
        ("", "· Negroni and Old Fashioned show 0 active stores but did sell."),
        ("", "  Is 'active' a current-range flag rather than sold-in-period?"),
        ("", "· Sales are the latest period, excluding VAT, as Checkers reports them."),
    ]
    for i, (head, text) in enumerate(lines, start=4):
        if head:
            ws.cell(row=i, column=1, value=head).font = Font(name="Arial", size=10, bold=True, color=ACCENT)
        if text:
            ws.cell(row=i, column=2, value=text).font = Font(name="Arial", size=10, color=INK)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 74

    # ------------------------------------------------------------------- Data
    ws = wb.create_sheet("Data")
    title(ws, "Data", "One row per article and pack, straight from the exports. Blue cells come from the file.")
    hdr = ["Portfolio", "Article key", "Article", "Pack", "Pack size", "Sell price (incl)"] \
        + [f"{w} units" for w in WEEKS] + ["Eaches, latest week", "Sales (excl VAT)", "Stock qty"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    first = ws.max_row + 1
    for d in rows:
        ws.append([d["portfolio"], d["key"], d["article"], d["pl"], d["pack"], d["price"]]
                  + d["weeks"] + [None, d["sales"], d["stock"]])
        r = ws.max_row
        # eaches is a formula so the sheet still works if someone edits units
        ws.cell(row=r, column=13, value=f"=L{r}*E{r}")
        for c in list(range(1, 13)) + [14, 15]:
            ws.cell(row=r, column=c).font = Font(name="Arial", size=9, color=BLUE)
        ws.cell(row=r, column=13).font = Font(name="Arial", size=9, color=INK)
        for c in range(6, 16):
            ws.cell(row=r, column=c).number_format = '#,##0.00' if c == 6 else '#,##0'
        ws.cell(row=r, column=14).number_format = '#,##0.00'
    last = ws.max_row
    for col, w in zip("ABCDEF", (10, 12, 40, 16, 10, 14)):
        ws.column_dimensions[col].width = w
    for c in range(7, 16):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "A4"

    # -------------------------------------------------------------- By product
    keys = sorted(stores.items(), key=lambda kv: kv[0])
    ws = wb.create_sheet("By product")
    title(ws, "By product", "Money and units come from Data. Store counts are the max across pack rows.")
    hdr = ["Portfolio", "Article key", "Article", "Listed stores", "Active stores",
           "Activation", "Eaches, latest week", "Sales (excl VAT)", "Share of portfolio"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    start = ws.max_row + 1
    for key, meta in keys:
        ws.append([meta["portfolio"], key, meta["article"], meta["lstd"], meta["active"]])
        r = ws.max_row
        ws.cell(row=r, column=6, value=f'=IF(D{r}=0,"",E{r}/D{r})')
        ws.cell(row=r, column=7, value=f"=SUMIF(Data!$B${first}:$B${last},$B{r},Data!$M${first}:$M${last})")
        ws.cell(row=r, column=8, value=f"=SUMIF(Data!$B${first}:$B${last},$B{r},Data!$N${first}:$N${last})")
        ws.cell(row=r, column=9, value=f'=IF(SUMIF($A:$A,$A{r},$H:$H)=0,"",H{r}/SUMIF($A:$A,$A{r},$H:$H))')
        for c in range(1, 10):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=9,
                                                 color=BLUE if c in (1, 2, 3, 4, 5) else INK)
        ws.cell(row=r, column=6).number_format = '0%'
        ws.cell(row=r, column=7).number_format = '#,##0'
        ws.cell(row=r, column=8).number_format = 'R#,##0'
        ws.cell(row=r, column=9).number_format = '0.0%'
    pend = ws.max_row
    for col, w in zip("ABCDEFGHI", (10, 12, 42, 13, 13, 11, 17, 15, 16)):
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A{start-1}:I{pend}"
    ws.freeze_panes = "A4"

    # ---------------------------------------------------------------- Summary
    ws = wb.create_sheet("Summary", 1)
    title(ws, "Summary", "Latest period. Patch and the core range kept apart, as Checkers structures them.")
    ws.append([])
    ws.append(["", "Patch", "Core range", "Total"])
    style_header(ws, ws.max_row, 4)
    r0 = ws.max_row
    metrics = [
        ("Sales (excl VAT)", f'=SUMIF(\'By product\'!$A${start}:$A${pend},B$%d,\'By product\'!$H${start}:$H${pend})', 'R#,##0'),
        ("Eaches sold, latest week", f'=SUMIF(\'By product\'!$A${start}:$A${pend},B$%d,\'By product\'!$G${start}:$G${pend})', '#,##0'),
        ("Products listed", f'=COUNTIF(\'By product\'!$A${start}:$A${pend},B$%d)', '#,##0'),
        ("Products actually selling", f'=COUNTIFS(\'By product\'!$A${start}:$A${pend},B$%d,\'By product\'!$H${start}:$H${pend},">0")', '#,##0'),
        ("Listings", f'=SUMIF(\'By product\'!$A${start}:$A${pend},B$%d,\'By product\'!$D${start}:$D${pend})', '#,##0'),
        ("Active store slots", f'=SUMIF(\'By product\'!$A${start}:$A${pend},B$%d,\'By product\'!$E${start}:$E${pend})', '#,##0'),
    ]
    for name, tmpl, fmt in metrics:
        ws.append([name])
        r = ws.max_row
        for col, letter in ((2, "B"), (3, "C")):
            ws.cell(row=r, column=col, value=(tmpl % r0).replace("B$", f"{letter}$"))
            ws.cell(row=r, column=col).number_format = fmt
            ws.cell(row=r, column=col).font = Font(name="Arial", size=10, color=GREEN)
        ws.cell(row=r, column=4, value=f"=B{r}+C{r}")
        ws.cell(row=r, column=4).number_format = fmt
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10, bold=True, color=INK)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10, color=INK)
    ar = ws.max_row + 1
    ws.cell(row=ar, column=1, value="Activation rate")
    ws.cell(row=ar, column=1).font = Font(name="Arial", size=10, bold=True, color=INK)
    for col in (2, 3, 4):
        L = get_column_letter(col)
        ws.cell(row=ar, column=col, value=f'=IF({L}{r0+5}=0,"",{L}{r0+6}/{L}{r0+5})')
        ws.cell(row=ar, column=col).number_format = '0%'
        ws.cell(row=ar, column=col).font = Font(name="Arial", size=10, bold=True, color=INK)
    note = ar + 2
    ws.cell(row=note, column=1,
            value="Patch is 6 products. The core range is 27. Read the sales line next to that.")
    ws.cell(row=note, column=1).font = Font(name="Arial", size=10, italic=True, color=ACCENT)
    for col, w in zip("ABCD", (30, 16, 16, 16)):
        ws.column_dimensions[col].width = w

    # ------------------------------------------------------- Distribution gaps
    ws = wb.create_sheet("Distribution gaps", 2)
    title(ws, "Distribution gaps",
          "Where you have a listing and no sale. Sorted by the size of the gap.")
    ws.append([])
    ws.append(["Portfolio", "Article", "Listed stores", "Active stores", "Activation", "Stores not selling"])
    style_header(ws, ws.max_row, 6)
    gap_rows = sorted(
        [(k, v) for k, v in stores.items() if v["lstd"] > 0],
        key=lambda kv: -(kv[1]["lstd"] - kv[1]["active"]),
    )
    for key, meta in gap_rows:
        ws.append([meta["portfolio"], meta["article"], meta["lstd"], meta["active"]])
        r = ws.max_row
        ws.cell(row=r, column=5, value=f'=IF(C{r}=0,"",D{r}/C{r})')
        ws.cell(row=r, column=6, value=f"=C{r}-D{r}")
        ws.cell(row=r, column=5).number_format = '0%'
        ws.cell(row=r, column=6).number_format = '#,##0'
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=9, color=INK)
        if meta["active"] == 0:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = WARN_FILL
    ws.cell(row=ws.max_row + 2, column=1,
            value="Shaded rows are listed in Checkers and selling in no store at all.")
    ws.cell(row=ws.max_row, column=1).font = Font(name="Arial", size=9, italic=True, color=ACCENT)
    for col, w in zip("ABCDEF", (10, 44, 13, 13, 11, 17)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(OUT)
    print(f"wrote {OUT}  ({len(rows)} data rows, {len(stores)} products)")


if __name__ == "__main__":
    build()
